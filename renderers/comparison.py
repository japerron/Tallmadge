"""
renderers/comparison.py — Comparison report generator
Produces an Excel workbook comparing two results envelopes.
"""

import re
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_BASE = Path(__file__).parent.parent / "output"


def _hex_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# Desired tab order for per-field sheets (fields not listed appear after, in diff order)
_SHEET_ORDER = [
    "Data Rows",
    "Data Sample (10 rows)",
    "Data Headers",
    "SQL",
    "WHERE Tokens",
    "Response Time",
    "Attributes Used",
    "Metrics Used",
]

CHANGE_COLORS = {
    "changed":        ("FFF9C4", "F57F17"),   # yellow bg, dark amber text
    "added":          ("E8F5E9", "1B6B2F"),   # green
    "removed":        ("FDECEA", "B71C1C"),   # red
    "new_prompt":     ("E3F2FD", "1565C0"),   # blue
    "missing_prompt": ("FCE4EC", "880E4F"),   # pink
}


def render(baseline: dict, current: dict, diffs: list, style: dict,
           out_dir: Path | None = None) -> Path:
    """
    Generate a comparison Excel workbook.
    diffs: output of core.results.compare()
    Returns path to saved .xlsx file.
    """
    from datetime import datetime
    run_date  = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    font_name = style["font"]
    primary   = style["primary"]

    if out_dir is None:
        out_dir = OUTPUT_BASE
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"tallmadge_baseline_{run_date}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    base_date = baseline.get("meta", {}).get("runDate", "baseline")[:10]
    curr_date = current.get("meta", {}).get("runDate", "current")[:10]

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")

    ws.merge_cells("A1:F1")
    ws["A1"] = "Tallmadge Comparison Report"
    ws["A1"].font      = Font(name=font_name, bold=True, size=16, color="FFFFFF")
    ws["A1"].fill      = _hex_fill(primary)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Baseline: {base_date}  ({baseline.get('meta',{}).get('mode','?')} mode)    vs    Current: {curr_date}  ({current.get('meta',{}).get('mode','?')} mode)"
    ws["A2"].font      = Font(name=font_name, size=10, color="FFFFFF")
    ws["A2"].fill      = _hex_fill("555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Stats
    change_types = {}
    for d in diffs:
        ct = d["change_type"]
        change_types[ct] = change_types.get(ct, 0) + 1

    row = 4
    ws[f"A{row}"] = "Change Type"
    ws[f"B{row}"] = "Count"
    for cell in [ws[f"A{row}"], ws[f"B{row}"]]:
        cell.font = Font(name=font_name, bold=True, size=10, color="FFFFFF")
        cell.fill = _hex_fill("444444")
        cell.border = _border()
    row += 1

    for ct, count in change_types.items():
        bg, fg = CHANGE_COLORS.get(ct, ("FFFFFF", "000000"))
        ws[f"A{row}"] = ct.replace("_", " ").title()
        ws[f"B{row}"] = count
        for col in "AB":
            ws[f"{col}{row}"].font   = Font(name=font_name, size=9, bold=(col=="A"), color=fg)
            ws[f"{col}{row}"].fill   = _hex_fill(bg)
            ws[f"{col}{row}"].border = _border()
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 10

    # ── All Changes sheet ───────────────────────────────────────────────────────
    wd = wb.create_sheet("All Changes")

    headers = ["Prompt #", "Category", "Prompt", "Field", "Change Type", "Baseline Value", "Current Value"]
    for ci, h in enumerate(headers, 1):
        cell = wd.cell(row=1, column=ci, value=h)
        cell.font      = Font(name=font_name, bold=True, size=9, color="FFFFFF")
        cell.fill      = _hex_fill("444444")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = _border()
    wd.row_dimensions[1].height = 18

    for ri, d in enumerate(diffs, 2):
        bg, fg = CHANGE_COLORS.get(d["change_type"], ("FFFFFF", "000000"))
        bfmt = _format_val(d.get("baseline_val"))
        cfmt = _format_val(d.get("current_val"))
        values = [
            d["id"],
            d["category"],
            d["prompt"][:80],
            d["field"],
            d["change_type"].replace("_", " ").title(),
            bfmt,
            cfmt,
        ]
        for ci, val in enumerate(values, 1):
            cell = wd.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name=font_name, size=9, color=fg if ci == 5 else "333333")
            cell.fill      = _hex_fill(bg if ci == 5 else ("FFFFFF" if ri % 2 == 0 else "F9F9F9"))
            cell.border    = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        wd.row_dimensions[ri].height = _row_height(bfmt, cfmt)

    col_widths = [8, 18, 45, 20, 16, 45, 45]
    for ci, w in enumerate(col_widths, 1):
        wd.column_dimensions[get_column_letter(ci)].width = w
    wd.freeze_panes = "A2"

    # ── Per-field sheets (in defined order, then any unlisted fields) ──────────
    fields_seen = {}
    for d in diffs:
        f = d["field"]
        if f == "*":
            continue   # new_prompt / missing_prompt — covered by All Changes only
        if f not in fields_seen:
            fields_seen[f] = []
        fields_seen[f].append(d)

    ordered_fields = [f for f in _SHEET_ORDER if f in fields_seen]
    ordered_fields += [f for f in fields_seen if f not in _SHEET_ORDER]

    def _write_field_sheet(field_name, field_diffs):
        safe_name = re.sub(r'[\\/*?\[\]:]', '-', field_name)[:31]
        wf = wb.create_sheet(safe_name)

        wf.merge_cells("A1:D1")
        wf["A1"] = f"Changes: {field_name}"
        wf["A1"].font      = Font(name=font_name, bold=True, size=12, color="FFFFFF")
        wf["A1"].fill      = _hex_fill(primary)
        wf["A1"].alignment = Alignment(horizontal="center")
        wf.row_dimensions[1].height = 24

        hdrs = ["Prompt #", "Prompt", "Baseline", "Current"]
        for ci, h in enumerate(hdrs, 1):
            cell = wf.cell(row=2, column=ci, value=h)
            cell.font   = Font(name=font_name, bold=True, size=9, color="FFFFFF")
            cell.fill   = _hex_fill("555555")
            cell.border = _border()

        for ri, d in enumerate(field_diffs, 3):
            bfmt = _format_val(d.get("baseline_val"))
            cfmt = _format_val(d.get("current_val"))
            vals = [d["id"], d["prompt"][:60], bfmt, cfmt]
            for ci, val in enumerate(vals, 1):
                cell = wf.cell(row=ri, column=ci, value=val)
                cell.font      = Font(name=font_name, size=9)
                cell.fill      = _hex_fill("FFFFFF" if ri % 2 == 0 else "F9F9F9")
                cell.border    = _border()
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            wf.row_dimensions[ri].height = _row_height(bfmt, cfmt)

        for ci, w in enumerate([8, 50, 40, 40], 1):
            wf.column_dimensions[get_column_letter(ci)].width = w
        wf.freeze_panes = "A3"

    for field_name in ordered_fields:
        _write_field_sheet(field_name, fields_seen[field_name])

    wb.save(str(out_path))
    return out_path


def _format_val(val, max_chars: int = 2000) -> str:
    if val is None:
        return "(none)"
    if isinstance(val, list):
        if not val:
            return "(empty)"
        if isinstance(val[0], list):
            # Tabular data rows (e.g. gridData_sample) — one row per line, cells pipe-separated
            lines = [" | ".join(str(c) for c in row) for row in val]
            return "\n".join(lines)[:max_chars]
        # List of scalars — newline-separate if items are long (e.g. SQL), comma-join if short
        # Threshold is 150 chars: column/attribute names stay comma-joined; SQL statements
        # (typically much longer) get newline-separated for readability
        if any(len(str(v)) > 150 for v in val):
            return "\n\n".join(str(v) for v in val)[:max_chars]
        return ", ".join(str(v) for v in val)[:max_chars]
    return str(val)[:max_chars]


def _row_height(*formatted_vals: str, line_height: int = 13, min_h: int = 20, max_h: int = 300) -> int:
    """Compute row height from the tallest formatted cell value."""
    max_lines = max((v.count("\n") + 1) for v in formatted_vals if v) if formatted_vals else 1
    return max(min_h, min(max_lines * line_height, max_h))

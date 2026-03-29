"""
renderers/xlsx.py — Excel workbook generator
Creates a formatted .xlsx with a summary sheet + one sheet per prompt.
"""

import re
import io
import base64
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
try:
    from openpyxl.drawing.image import Image as XLImage
    _HAS_XLIMAGE = True
except ImportError:
    _HAS_XLIMAGE = False

from core.results import parse_grid_data, format_sql, strip_markdown
from core.color import hex_darken, hex_lighten

OUTPUT_BASE = Path(__file__).parent.parent / "output"


def _hex_fill(hex_color: str) -> PatternFill:
    c = hex_color.lstrip("#")
    return PatternFill("solid", fgColor=c)


def _border(style="thin") -> Border:
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names max 31 chars, no special chars."""
    name = re.sub(r'[\\/*?:\[\]]', '', name)
    return name[:31]


def _add_wide_view_sheet(
    wb,
    results: list,
    font_name: str,
    primary: str,        # hex WITHOUT leading #
    primary_light: str,  # hex WITHOUT leading # — for alternating row tint
) -> None:
    """
    Insert a 'Wide View' sheet (as second sheet, after Summary).
    One row per prompt; all fields spread across columns.
    SQL is formatted with keyword newlines; Explanation has Markdown stripped.
    """
    ws = wb.create_sheet("Wide View", 1)   # position: after Summary

    col_defs = [
        ("#",              5),
        ("Category",      18),
        ("Prompt",        45),
        ("Status",        10),
        ("RT (s)",         8),
        ("Cache",          8),
        ("Response Text", 42),
        ("Interpretation",35),
        ("Explanation",   40),
        ("Insights",      35),
        ("SQL",           52),
        ("WHERE Tokens",  30),
        ("Attributes Used",30),
        ("Metrics Used",  30),
        ("Datasets Used", 25),
        ("Answer Type",   12),
        ("Data Rows",     10),
    ]
    headers    = [d[0] for d in col_defs]
    col_widths = [d[1] for d in col_defs]
    SQL_COL    = headers.index("SQL") + 1   # 1-based

    # ── Header row ─────────────────────────────────────────────────────────────
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(name=font_name, bold=True, size=9, color="FFFFFF")
        cell.fill      = _hex_fill(primary)
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border    = _border()
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Data rows ──────────────────────────────────────────────────────────────
    for row_i, r in enumerate(results, 2):
        is_ok = r["status"] == "Success"
        bg    = "FFFFFF" if row_i % 2 == 0 else primary_light

        sql_raw  = r.get("sqlQueries") or []
        sql_text = "\n\n".join(format_sql(q) for q in sql_raw) if sql_raw else ""

        cache_val = r.get("isCacheUsed")
        cache_str = ("Yes" if cache_val else "No") if cache_val is not None else ""

        grid_rows = r.get("gridData") or []
        values = [
            r["id"],
            r["category"],
            r["prompt"],
            r["status"],
            r.get("responseTime"),
            cache_str,
            r.get("responseText") or "",
            r.get("interpretedQuestion") or "",
            strip_markdown(r.get("explanation") or ""),
            r.get("insights") or "",
            sql_text,
            ", ".join(r.get("whereClauseTokens") or []),
            ", ".join(r.get("attributesUsed") or []),
            ", ".join(r.get("metricsUsed")    or []),
            ", ".join(r.get("datasetsUsed")   or []),
            r.get("answerType") or "",
            len(grid_rows) if grid_rows else None,
        ]

        for col_i, val in enumerate(values, 1):
            is_sql    = (col_i == SQL_COL)
            is_status = (col_i == 4)
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font = Font(
                name  = "Courier New" if is_sql else font_name,
                size  = 8             if is_sql else 9,
                bold  = is_status,
                color = ("1B6B2F" if (is_status and is_ok)
                         else "B71C1C" if (is_status and not is_ok)
                         else "333333"),
            )
            cell.fill      = _hex_fill(bg)
            cell.border    = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[row_i].height = 60   # tall rows for multi-line content

    # ── Column widths ──────────────────────────────────────────────────────────
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _add_data_sheets(wb, results: list, font_name: str, primary: str,
                     primary_dark: str, primary_light: str) -> None:
    """
    Add one 'Data-{id}' sheet per prompt that has gridData rows.
    Used in 'wide' layout where no detail sheets are produced.
    """
    for r in results:
        grid_data = r.get("gridData") or []
        if not grid_data:
            continue
        headers, rows = parse_grid_data(grid_data)
        if not headers or not rows:
            continue

        sheet_name = _safe_sheet_name(f"Data-{r['id']}")
        ws = wb.create_sheet(sheet_name)

        # Title row
        n_cols = len(headers)
        end_col = get_column_letter(max(n_cols, 1))
        ws.merge_cells(f"A1:{end_col}1")
        ws["A1"] = f"P{r['id']} — {r['category']}  |  {r['prompt'][:60]}"
        ws["A1"].font      = Font(name=font_name, bold=True, size=10, color="FFFFFF")
        ws["A1"].fill      = _hex_fill(primary)
        ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[1].height = 22

        # Header row
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=str(h))
            cell.font      = Font(name=font_name, bold=True, size=9, color="FFFFFF")
            cell.fill      = _hex_fill(primary_dark)
            cell.border    = _border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 16
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"

        # Data rows
        for ri, row in enumerate(rows, 3):
            bg = "FFFFFF" if ri % 2 == 1 else primary_light
            vals = row if isinstance(row, (list, tuple)) else list(row.values())
            for ci, val in enumerate(vals[:n_cols], 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = Font(name=font_name, size=9)
                cell.fill      = _hex_fill(bg)
                cell.border    = _border()
                cell.alignment = Alignment(vertical="top")
            ws.row_dimensions[ri].height = 14

        # Auto-size columns (capped)
        for ci, h in enumerate(headers, 1):
            col_vals = [str(h)] + [
                str((row if isinstance(row, (list, tuple)) else list(row.values()))[ci - 1] or "")
                for row in rows
            ]
            width = min(max(len(v) for v in col_vals) + 2, 40)
            ws.column_dimensions[get_column_letter(ci)].width = width


def render(envelope: dict, style: dict, out_dir: Path | None = None,
           layout: str = "detail") -> Path:
    """
    Generate an Excel workbook from results envelope.

    layout options:
      "detail"  (default) — Summary + one sheet per prompt
      "wide"              — Summary + Wide View sheet (one row per prompt, all fields as columns)
      "both"              — Summary + Wide View + one sheet per prompt

    Returns the path to the saved .xlsx file.
    """
    results  = envelope.get("results", [])
    meta     = envelope.get("meta", {})
    _raw_date = meta.get("runDate", "")
    run_date  = _raw_date[:10]                                        # YYYY-MM-DD  (display)
    run_ts    = (_raw_date[:19].replace("T", "_").replace(":", "-")   # YYYY-MM-DD_HH-MM-SS
                 if len(_raw_date) >= 19 else run_date)               # (filename)
    mode     = meta.get("mode", "standard")

    if out_dir is None:
        out_dir = OUTPUT_BASE
    out_dir.mkdir(parents=True, exist_ok=True)

    out_path = out_dir / f"tallmadge_{run_ts}.xlsx"

    primary   = style["primary"].lstrip("#")
    secondary = style["secondary"].lstrip("#")
    accent    = style["accent"].lstrip("#")
    font_name = style["font"]

    # Derived brand shades — computed once, used throughout the workbook
    primary_dark  = hex_darken(style["primary"],  0.78).lstrip("#")  # for column/section headers
    primary_light = hex_lighten(style["primary"], 0.88).lstrip("#")  # for alternating data rows

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = "Agent Test Results — Summary"
    ws["A1"].font      = Font(name=font_name, bold=True, size=16, color="FFFFFF")
    ws["A1"].fill      = _hex_fill(style["primary"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Meta info
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Run Date: {run_date}  |  Mode: {mode.title()}  |  Prompts: {meta.get('totalPrompts',0)}  |  Success: {meta.get('successful',0)}  |  Errors: {meta.get('errors',0)}"
    ws["A2"].font      = Font(name=font_name, size=10, color="FFFFFF")
    ws["A2"].fill      = _hex_fill(style["secondary"])
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Header row
    headers = ["#", "Category", "Prompt", "Status", "Response Time (s)", "Response", "Attributes Used", "Metrics Used", "Data Rows"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(name=font_name, bold=True, size=10, color="FFFFFF")
        cell.fill      = _hex_fill(primary_dark)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = _border()
    ws.row_dimensions[3].height = 20

    # Data rows
    for row_i, r in enumerate(results, 4):
        is_ok = r["status"] == "Success"
        bg    = "FFFFFF" if row_i % 2 == 0 else primary_light

        _gd = r.get("gridData") or []
        values = [
            r["id"],
            r["category"],
            r["prompt"],
            r["status"],
            r.get("responseTime"),
            r.get("responseText") or "",
            ", ".join(r.get("attributesUsed") or []),
            ", ".join(r.get("metricsUsed") or []),
            len(_gd) if _gd else None,
        ]
        for col_i, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font   = Font(name=font_name, size=9)
            cell.fill   = _hex_fill(bg)
            cell.border = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_i == 4:  # Status column
                cell.font = Font(
                    name=font_name, size=9, bold=True,
                    color="1B6B2F" if is_ok else "B71C1C"
                )

        # Hyperlink — to Wide View row (always), or to detail sheet when it will exist
        if layout in ("wide", "both"):
            # Row in Wide View = row_i - 4 (data starts at row 2 there) + 1 = row_i - 3
            wv_row = row_i - 4 + 2   # Summary data starts at row 4; Wide View data at row 2
            ws.cell(row=row_i, column=1).hyperlink = f"#'Wide View'!A{wv_row}"
            ws.cell(row=row_i, column=1).font = Font(
                name=font_name, size=9, color=primary, underline="single"
            )
        if layout in ("detail", "both"):
            sheet_ref = f"P{r['id']}"
            ws.cell(row=row_i, column=1).hyperlink = f"#{sheet_ref}!A1"
            ws.cell(row=row_i, column=1).font = Font(
                name=font_name, size=9, color=primary, underline="single"
            )
        ws.row_dimensions[row_i].height = 40

    # Column widths: #, Category, Prompt, Status, RT, Response, Attrs, Metrics, Data Rows
    col_widths = [5, 18, 60, 10, 16, 45, 35, 45, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    # ── Wide View sheet (layout: "wide" or "both") ──────────────────────────────
    if layout in ("wide", "both"):
        _add_wide_view_sheet(wb, results, font_name, primary, primary_light)

    # ── Data sheets (one per prompt with gridData; "wide" layout only) ──────────
    # In "detail"/"both" layouts the data already appears in detail sheets.
    if layout == "wide":
        _add_data_sheets(wb, results, font_name, primary, primary_dark, primary_light)

    # ── Detail sheets (one per prompt; layout: "detail" or "both") ─────────────
    if layout not in ("detail", "both"):
        wb.save(str(out_path))
        return out_path

    for r in results:
        sheet_name = f"P{r['id']}"
        ws2 = wb.create_sheet(sheet_name)

        def _write_section(row: int, label: str, value, wide: bool = True,
                           mono: bool = False) -> int:
            """Write a label + value section. Returns next row."""
            if value is None:
                return row
            # Label
            ws2.merge_cells(f"A{row}:F{row}")
            ws2[f"A{row}"] = label
            ws2[f"A{row}"].font = Font(name=font_name, bold=True, size=9, color="FFFFFF")
            ws2[f"A{row}"].fill = _hex_fill(primary_dark)
            ws2[f"A{row}"].alignment = Alignment(vertical="center")
            ws2.row_dimensions[row].height = 18
            row += 1

            # Value
            display = value
            if isinstance(value, list):
                display = "\n".join(str(v) for v in value)
            ws2.merge_cells(f"A{row}:F{row + (3 if wide else 0)}")
            cell = ws2[f"A{row}"]
            cell.value     = display
            cell.font      = Font(name="Courier New" if mono else font_name, size=8 if mono else 9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill      = _hex_fill("FAFAFA")
            cell.border    = _border()
            end_row = row + (4 if wide else 1)
            for r2 in range(row, end_row):
                ws2.row_dimensions[r2].height = 15
            return end_row + 1

        # Title
        ws2.merge_cells("A1:F1")
        ws2["A1"] = f"Prompt {r['id']} — {r['category']}"
        ws2["A1"].font      = Font(name=font_name, bold=True, size=13, color="FFFFFF")
        ws2["A1"].fill      = _hex_fill(style["primary"])
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 28

        # Prompt text
        ws2.merge_cells("A2:F2")
        ws2["A2"] = r["prompt"]
        ws2["A2"].font      = Font(name=font_name, italic=True, size=10)
        ws2["A2"].alignment = Alignment(wrap_text=True)
        ws2["A2"].fill      = _hex_fill("F0F0F0")
        ws2.row_dimensions[2].height = 30

        # Meta row
        ws2["A3"] = "Status"
        ws2["B3"] = r.get("status", "")
        ws2["C3"] = "Response Time"
        ws2["D3"] = f"{r.get('responseTime', '')}s"
        ws2["E3"] = "Mode"
        ws2["F3"] = r.get("mode", "")
        for col in "ABCDEF":
            ws2[f"{col}3"].font   = Font(name=font_name, size=9, bold=(col in "ACE"))
            ws2[f"{col}3"].fill   = _hex_fill("EEEEEE")
            ws2[f"{col}3"].border = _border()
        ws2.row_dimensions[3].height = 16

        # Pre-process text fields for readability
        explanation_text = strip_markdown(r.get("explanation"))
        sql_queries_raw  = r.get("sqlQueries") or []
        sql_formatted    = [format_sql(q) for q in sql_queries_raw] if sql_queries_raw else None

        cur_row = 5
        cur_row = _write_section(cur_row, "Response Text",     r.get("responseText"))
        cur_row = _write_section(cur_row, "Interpretation",    r.get("interpretedQuestion"))
        cur_row = _write_section(cur_row, "Explanation",       explanation_text or None)
        cur_row = _write_section(cur_row, "Insights",          r.get("insights"))
        cur_row = _write_section(cur_row, "SQL Queries",       sql_formatted, mono=True)
        cur_row = _write_section(cur_row, "WHERE Tokens",      ", ".join(r.get("whereClauseTokens") or []) or None, wide=False)
        cur_row = _write_section(cur_row, "Attributes Used",   ", ".join(r.get("attributesUsed") or []) or None, wide=False)
        cur_row = _write_section(cur_row, "Metrics Used",      ", ".join(r.get("metricsUsed") or []) or None, wide=False)
        cur_row = _write_section(cur_row, "Attribute Forms",   ", ".join(r.get("attributeFormsUsed") or []) or None, wide=False)
        cur_row = _write_section(cur_row, "Datasets Used",     ", ".join(r.get("datasetsUsed") or []) or None, wide=False)

        # ── Grid data table (from a.data / gridData) ───────────────────────────
        g_headers, g_rows = parse_grid_data(r.get("gridData"))
        if g_headers and g_rows:
            ws2.merge_cells(f"A{cur_row}:F{cur_row}")
            ws2[f"A{cur_row}"] = "Data Grid"
            ws2[f"A{cur_row}"].font  = Font(name=font_name, bold=True, size=9, color="FFFFFF")
            ws2[f"A{cur_row}"].fill  = _hex_fill(primary_dark)
            ws2.row_dimensions[cur_row].height = 18
            cur_row += 1

            for ci, h in enumerate(g_headers, 1):
                cell = ws2.cell(row=cur_row, column=ci, value=str(h))
                cell.font   = Font(name=font_name, bold=True, size=9, color="FFFFFF")
                cell.fill   = _hex_fill(primary)
                cell.border = _border()
            ws2.row_dimensions[cur_row].height = 16
            cur_row += 1

            for dr_i, dr in enumerate(g_rows):
                bg = "FFFFFF" if dr_i % 2 == 0 else primary_light
                vals = dr if isinstance(dr, (list, tuple)) else list(dr.values())
                for ci, val in enumerate(vals, 1):
                    cell = ws2.cell(row=cur_row, column=ci, value=val)
                    cell.font   = Font(name=font_name, size=9)
                    cell.fill   = _hex_fill(bg)
                    cell.border = _border()
                ws2.row_dimensions[cur_row].height = 14
                cur_row += 1

            cur_row += 1  # gap after grid

        # ── Images ─────────────────────────────────────────────────────────────
        if _HAS_XLIMAGE:
            for img in (r.get("imagesData") or []):
                try:
                    img_bytes = base64.b64decode(img["data"])
                    xl_img = XLImage(io.BytesIO(img_bytes))
                    xl_img.width  = 600   # pixels
                    xl_img.height = 450
                    ws2.add_image(xl_img, f"A{cur_row}")
                    cur_row += 32   # ~450px ÷ ~14px per row
                except Exception:
                    pass

        # Chart data as table
        chart_data = r.get("chartData")
        if chart_data:
            charts = chart_data.get("charts") or ([chart_data] if isinstance(chart_data, dict) else [])
            if charts:
                chart = charts[0] if isinstance(charts, list) else charts
                rows  = chart.get("data", [])
                option = chart.get("option", {})
                columns = option.get("columns", [])
                if rows and columns:
                    # Section label
                    ws2.merge_cells(f"A{cur_row}:F{cur_row}")
                    ws2[f"A{cur_row}"] = "Chart Data"
                    ws2[f"A{cur_row}"].font  = Font(name=font_name, bold=True, size=9, color="FFFFFF")
                    ws2[f"A{cur_row}"].fill  = _hex_fill(primary_dark)
                    ws2.row_dimensions[cur_row].height = 18
                    cur_row += 1

                    # Header
                    for ci, col in enumerate(columns, 1):
                        cell = ws2.cell(row=cur_row, column=ci, value=col["column_name"])
                        cell.font   = Font(name=font_name, bold=True, size=9, color="FFFFFF")
                        cell.fill   = _hex_fill(primary)
                        cell.border = _border()
                    ws2.row_dimensions[cur_row].height = 16
                    cur_row += 1

                    # Data rows
                    for dr_i, dr in enumerate(rows):
                        bg = "FFFFFF" if dr_i % 2 == 0 else primary_light
                        for ci, col in enumerate(columns, 1):
                            val  = dr.get(col["column_name"], "")
                            cell = ws2.cell(row=cur_row, column=ci, value=val)
                            cell.font   = Font(name=font_name, size=9)
                            cell.fill   = _hex_fill(bg)
                            cell.border = _border()
                            if col.get("type") == 2:
                                cell.alignment = Alignment(horizontal="right")
                        ws2.row_dimensions[cur_row].height = 14
                        cur_row += 1

        for ci in range(1, 7):
            ws2.column_dimensions[get_column_letter(ci)].width = 30

    wb.save(str(out_path))
    return out_path

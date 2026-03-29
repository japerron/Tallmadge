"""
renderers/expected_report.py — Expected-values comparison report

Generates a scored Excel workbook comparing actual results against a standard.

Sheet layout (mirrors comparison.py conventions):
  Score Card    — title, score-distribution summary, per-prompt scored table
  All Findings  — every deduction row (field, expected, actual, δ, details)
  WHERE Tokens  — per-field drill-down for WHERE mismatches
  Attributes Used
  Metrics Used
  Data Rows
  SQL Comparison — LLM verdict per prompt (only when sql_judgment data is present)
"""

import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.results import format_sql

OUTPUT_BASE = Path(__file__).parent.parent / "output"

# ── Score colour bands (lo inclusive, hi inclusive) ───────────────────────────
_SCORE_BANDS = [
    (5.0,  5.0,  "C8E6C9", "1B5E20"),   # perfect  — green
    (4.0,  4.99, "FFF9C4", "F57F17"),   # good     — amber
    (3.0,  3.99, "FFE0B2", "E65100"),   # fair     — orange
    (0.0,  2.99, "FDECEA", "B71C1C"),   # poor     — red
]
_UNMATCHED_BG, _UNMATCHED_FG = "F5F5F5", "757575"
_INFO_BG,      _INFO_FG      = "E3F2FD", "1565C0"   # prompt-text mismatch rows

_FIELD_ORDER = [
    "WHERE Tokens", "HAVING Tokens",
    "Attributes Used", "Metrics Used", "Other Used",
    "Data Rows",
]

# Maps each field to: (standard_key, fmt_expected, result_key, fmt_actual)
# result_key is looked up on the result dict (res); for derived fields (_havingTokens,
# _otherFound) it is stored there by score_results() at scoring time.
_FIELD_DISPLAY = {
    "WHERE Tokens": (
        "where_map",
        lambda v: ", ".join(sorted(v.values())) if v else "—",
        "whereClauseTokens",
        lambda v: ", ".join(sorted(v)) if v else "(none)",
    ),
    "HAVING Tokens": (
        "having_map",
        lambda v: ", ".join(sorted(v.values())) if v else "—",
        "_havingTokens",
        lambda v: ", ".join(sorted(v)) if v else "(none)",
    ),
    "Attributes Used": (
        "attrs",
        lambda v: ", ".join(sorted(v, key=str.lower)) if v else "—",
        "attributesUsed",
        lambda v: ", ".join(sorted(v, key=str.lower)) if v else "(none)",
    ),
    "Metrics Used": (
        "metrics",
        lambda v: ", ".join(sorted(v, key=str.lower)) if v else "—",
        "metricsUsed",
        lambda v: ", ".join(sorted(v, key=str.lower)) if v else "(none)",
    ),
    "Other Used": (
        "other",
        lambda v: ", ".join(sorted(v, key=str.lower)) if v else "—",
        "_otherFound",
        lambda v: ", ".join(sorted(v, key=str.lower)) if v else "(none)",
    ),
    "Data Rows": (
        "data_rows",
        lambda v: str(v) if v is not None else "—",
        "gridData",
        lambda v: str(len(v)) if isinstance(v, list) else "(none)",
    ),
}

_MATCH_BG, _MATCH_FG = "E8F5E9", "2E7D32"   # light green — field passed
_NA_BG,    _NA_FG    = "F5F5F5", "9E9E9E"   # grey — field not in standard

# Fields whose sheets include an "Optional" column (between Expected and Actual)
_OPTIONAL_STD_KEY = {
    "Attributes Used": "optional_attrs",
    "Metrics Used":    "optional_metrics",
    "Other Used":      "optional_other",
}

# Safe sheet name for each field (Excel sheet-name constraints applied)
_FIELD_SHEET_NAME: dict[str, str] = {
    f: re.sub(r'[\\/*?\[\]:]', '-', f)[:31]
    for f in _FIELD_ORDER
}

# Column index of "Final Deduction" in each field sheet.
# Fields with an Optional column have one extra col, pushing Final Deduction to col 7.
_FIELD_SHEET_FD_COL: dict[str, int] = {
    f: (7 if f in _OPTIONAL_STD_KEY else 6)
    for f in _FIELD_ORDER
}


def _delta_lookup(row: int, field: str) -> str:
    """Score Card delta cell formula: look up Final Deduction from the field sheet.
    Returns blank when the value is 0 (no deduction / field not in standard)."""
    sheet   = _FIELD_SHEET_NAME[field]
    col_idx = _FIELD_SHEET_FD_COL[field]
    col_ext = get_column_letter(col_idx + 1)
    lkp = f"VLOOKUP(A{row},'{sheet}'!$A:${col_ext},{col_idx},FALSE)"
    return f'=IF(IFERROR({lkp},0)>0,IFERROR({lkp},0),"")'


def _field_score_expr(row: int) -> str:
    """Field Score formula: MAX(0, 5 − sum of all Final Deduction VLOOKUPs)."""
    parts = []
    for f in _FIELD_ORDER:
        sheet   = _FIELD_SHEET_NAME[f]
        col_idx = _FIELD_SHEET_FD_COL[f]
        col_ext = get_column_letter(col_idx + 1)
        parts.append(
            f"IFERROR(VLOOKUP(A{row},'{sheet}'!$A:${col_ext},{col_idx},FALSE),0)"
        )
    return "=MAX(0,5-" + "-".join(parts) + ")"



# ── Style helpers ─────────────────────────────────────────────────────────────

def _score_colors(score) -> tuple[str, str]:
    if score is None:
        return _UNMATCHED_BG, _UNMATCHED_FG
    for lo, hi, bg, fg in _SCORE_BANDS:
        if lo <= round(score, 4) <= hi:
            return bg, fg
    return "FFFFFF", "000000"


def _hex_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))


def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, col, value, font_name, bg="444444", fg="FFFFFF",
         bold=True, size=9, wrap=False, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=font_name, bold=bold, size=size, color=fg)
    c.fill      = _hex_fill(bg)
    c.border    = _border()
    c.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical="center")
    return c


def _dat(ws, row, col, value, font_name, bg="FFFFFF", fg="333333",
         bold=False, size=9, wrap=False, align="left", valign="top"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=font_name, bold=bold, size=size, color=fg)
    c.fill      = _hex_fill(bg)
    c.border    = _border()
    c.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical=valign)
    return c


def _row_ht(*vals, lh=13, mn=18, mx=300) -> int:
    mx_lines = max((str(v).count("\n") + 1 for v in vals if v), default=1)
    return max(mn, min(mx_lines * lh, mx))


def _alt(ri) -> str:
    """Alternating row background."""
    return "FFFFFF" if ri % 2 == 0 else "F9F9F9"


# ── Public entry point ────────────────────────────────────────────────────────

def render(scored: list[dict], envelope: dict, style: dict,
           standard_path: str, standard: list[dict] | None = None,
           out_dir: Path | None = None) -> Path:
    """
    Generate the expected-values report workbook.
    scored       : output of core.expected.score_results()
    envelope     : the results envelope (for run metadata)
    style        : Tallmadge style dict (font, primary, …)
    standard_path: path to the source Excel standard file (for display)
    standard     : parsed standard records (from core.expected.load_standard);
                   when supplied, Match rows in field sheets show expected/actual values
    Returns the path to the saved .xlsx file.
    """
    run_date  = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    font_name = style["font"]
    primary   = style["primary"].lstrip("#")

    if out_dir is None:
        out_dir = OUTPUT_BASE
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"tallmadge_goldstandard_{run_date}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    std_name = Path(standard_path).name
    run_ts   = envelope.get("meta", {}).get("runDate", "")[:10]

    # Lookup maps for field-sheet Match rows
    std_map    = {r["prompt"].strip().lower(): r for r in (standard or [])}
    result_map = {r.get("id"): r for r in envelope.get("results", [])}

    _write_score_card(wb, scored, font_name, primary, run_ts, std_name)
    _write_all_findings(wb, scored, font_name, primary)
    _write_field_sheets(wb, scored, font_name, primary, std_map, result_map)
    _write_sql_sheet(wb, scored, font_name, primary, std_map, result_map)

    wb.save(str(out_path))
    return out_path


# ── Score Card sheet ──────────────────────────────────────────────────────────

def _write_score_card(wb, scored, font_name, primary, run_ts, std_name):
    ws = wb.create_sheet("Score Card")

    # ── Title / subtitle ──────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ws["A1"] = "Tallmadge Expected-Values Report"
    ws["A1"].font      = Font(name=font_name, bold=True, size=16, color="FFFFFF")
    ws["A1"].fill      = _hex_fill(primary)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Run: {run_ts}    Standard: {std_name}"
    ws["A2"].font      = Font(name=font_name, size=10, color="FFFFFF")
    ws["A2"].fill      = _hex_fill("555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # ── Score-distribution summary ────────────────────────────────────────────
    row = 4
    for ci, h in enumerate(["Score Band", "Count", "% of Scored"], 1):
        _hdr(ws, row, ci, h, font_name)
    row += 1

    matched   = [s for s in scored if s["score"] is not None]
    unmatched = [s for s in scored if s["score"] is None]

    # Track upper-table row numbers for formula overwrite after lower table is written
    upper_rows: dict[str, int] = {}

    bands = [
        ("perfect", "Perfect  (5.0)",     "C8E6C9", "1B5E20"),
        ("good",    "Good     (4.0–4.9)", "FFF9C4", "F57F17"),
        ("fair",    "Fair     (3.0–3.9)", "FFE0B2", "E65100"),
        ("poor",    "Poor     (< 3.0)",   "FDECEA", "B71C1C"),
    ]
    for key, label, bg, fg in bands:
        upper_rows[key] = row
        count = sum(1 for s in matched if {
            "perfect": lambda v: v == 5.0,
            "good":    lambda v: 4.0 <= v < 5.0,
            "fair":    lambda v: 3.0 <= v < 4.0,
            "poor":    lambda v: v < 3.0,
        }[key](s["score"]))
        pct = f"{count / len(matched) * 100:.0f}%" if matched else "—"
        for ci, val in enumerate([label, count, pct], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font      = Font(name=font_name, size=9, bold=(ci == 1), color=fg)
            c.fill      = _hex_fill(bg)
            c.border    = _border()
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center")
        row += 1

    if unmatched:
        upper_rows["unmatched"] = row
        for ci, val in enumerate(["Unmatched", "—", "—"], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font      = Font(name=font_name, size=9, bold=(ci == 1),
                               color=_UNMATCHED_FG)
            c.fill      = _hex_fill(_UNMATCHED_BG)
            c.border    = _border()
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center")
        row += 1

    if matched:
        upper_rows["field_avg"] = row
        avg = sum(s["score"] for s in matched) / len(matched)
        for ci, val in enumerate(["Field Score Average", round(avg, 2), ""], 1):
            c = ws.cell(row=row, column=ci, value=val if val != "" else None)
            c.font   = Font(name=font_name, size=9, bold=True)
            c.border = _border()
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center")
        row += 1

    sql_scored = [s for s in matched if s.get("sql_score") is not None]
    if sql_scored:
        upper_rows["sql_avg"] = row
        sql_avg = sum(s["sql_score"] for s in sql_scored) / len(sql_scored)
        for ci, val in enumerate(["SQL Score Average", round(sql_avg, 2), ""], 1):
            c = ws.cell(row=row, column=ci, value=val if val != "" else None)
            c.font   = Font(name=font_name, size=9, bold=True)
            c.border = _border()
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center")
        row += 1

    # ── Per-prompt scored table ───────────────────────────────────────────────
    row += 1  # blank separator
    prompt_hdr_row = row
    hdrs = ["#", "Category", "Prompt", "WHERE δ", "HAVING δ", "Attrs δ",
            "Metrics δ", "Other δ", "Rows δ", "Field Score", "SQL Score", "Notes"]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, row, ci, h, font_name)
    row += 1

    for s in scored:
        score_bg, score_fg = _score_colors(s["score"])

        notes = ""
        if not s["matched"]:
            notes = "No match in standard"
        elif s["prompt_differs"]:
            notes = "Prompt text differs"

        sql_j     = s.get("sql_judgment")
        sql_score = s.get("sql_score")
        if sql_j is None:
            sql_display = ""
            sql_formula = ""
        elif sql_score is None:
            sql_display = "err"
            sql_formula = (f'=IFERROR(VLOOKUP(A{row},'
                           f"'SQL Comparison'!$A:$D,4,FALSE),\"\")")
        else:
            sql_display = sql_score
            sql_formula = (f'=IFERROR(VLOOKUP(A{row},'
                           f"'SQL Comparison'!$A:$D,4,FALSE),\"\")")

        # Delta columns and Field Score: formulas for matched prompts, static for unmatched
        if s["matched"]:
            where_d   = _delta_lookup(row, "WHERE Tokens")
            having_d  = _delta_lookup(row, "HAVING Tokens")
            attrs_d   = _delta_lookup(row, "Attributes Used")
            metrics_d = _delta_lookup(row, "Metrics Used")
            other_d   = _delta_lookup(row, "Other Used")
            rows_d    = _delta_lookup(row, "Data Rows")
            field_score_val = _field_score_expr(row)
        else:
            where_d = having_d = attrs_d = metrics_d = other_d = rows_d = ""
            field_score_val = "—"

        vals = [
            s["id"],
            s["category"],
            s["prompt"],
            where_d,
            having_d,
            attrs_d,
            metrics_d,
            other_d,
            rows_d,
            field_score_val,   # col 10: Field Score
            sql_formula,       # col 11: SQL Score
            notes,
        ]
        for ci, val in enumerate(vals, 1):
            va = "top" if ci == 3 else "center"
            if ci == 10:   # Field Score — score-band colour from Python value
                _dat(ws, row, ci, val, font_name,
                     bg=score_bg, fg=score_fg, bold=True,
                     align="center", valign=va)
            elif ci == 11:   # SQL Score — formula; colour from Python value
                if sql_j and isinstance(sql_display, float):
                    sql_bg, sql_fg = _score_colors(sql_display)
                    _dat(ws, row, ci, val, font_name,
                         bg=sql_bg, fg=sql_fg, bold=True, align="center", valign=va)
                elif sql_j and sql_display == "err":
                    _dat(ws, row, ci, val, font_name,
                         bg=_UNMATCHED_BG, fg=_UNMATCHED_FG, align="center", valign=va)
                else:
                    _dat(ws, row, ci, val, font_name,
                         bg=_alt(row), align="center", valign=va)
            else:
                _dat(ws, row, ci, val, font_name, bg=_alt(row),
                     wrap=(ci == 3),
                     align="center" if ci in (1, 4, 5, 6, 7, 8, 9, 12) else "left",
                     valign=va)
        row += 1

    # ── Overwrite upper table with live formulas now that data range is known ─
    data_start = prompt_hdr_row + 1
    data_end   = prompt_hdr_row + len(scored)
    j_rng      = f"$J${data_start}:$J${data_end}"
    k_rng      = f"$K${data_start}:$K${data_end}"

    band_criteria = {
        "perfect": f"=COUNTIF({j_rng},5)",
        "good":    f'=COUNTIFS({j_rng},">=4",{j_rng},"<5")',
        "fair":    f'=COUNTIFS({j_rng},">=3",{j_rng},"<4")',
        "poor":    f'=COUNTIFS({j_rng},">=0",{j_rng},"<3")',
    }
    for key, formula in band_criteria.items():
        r = upper_rows.get(key)
        if r:
            ws.cell(row=r, column=2).value = formula
            ws.cell(row=r, column=3).value = (
                f'=IFERROR(TEXT(B{r}/COUNT({j_rng}),"0%"),"—")'
            )

    if "unmatched" in upper_rows:
        r = upper_rows["unmatched"]
        ws.cell(row=r, column=2).value = f"=COUNTA({j_rng})-COUNT({j_rng})"

    if "field_avg" in upper_rows:
        r = upper_rows["field_avg"]
        ws.cell(row=r, column=2).value = (
            f'=IFERROR(ROUND(AVERAGE({j_rng}),2),"")'
        )

    if "sql_avg" in upper_rows:
        r = upper_rows["sql_avg"]
        ws.cell(row=r, column=2).value = (
            f'=IFERROR(AVERAGEIF({k_rng},">=0"),"")'
        )

    for ci, w in enumerate([6, 22, 52, 10, 10, 10, 10, 10, 10, 10, 10, 18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{prompt_hdr_row + 1}"


# ── Deduction detail formatter ────────────────────────────────────────────────

def _fmt_deduction_detail(d: dict) -> str:
    """
    Format the Details cell text for a deduction record.
    Handles WHERE/HAVING paired diffs (substituted, op_diff) as well as the
    plain missing/extra lists used by Attributes, Metrics, Other, and Data Rows.
    """
    parts = []
    if d.get("missing"):
        parts.append("Missing: " + ", ".join(d["missing"]))
    for gold, actual in d.get("substituted", []):
        parts.append(f"Substituted: {gold}  →  {actual}")
    for gold, actual in d.get("op_diff", []):
        parts.append(f"Operator diff: {gold}  →  {actual}")
    if d.get("extra"):
        parts.append("Extra: " + ", ".join(d["extra"]))
    return "\n".join(parts) or "—"


# ── All Findings sheet ────────────────────────────────────────────────────────

def _write_all_findings(wb, scored, font_name, primary):
    ws = wb.create_sheet("All Findings")

    hdrs = ["#", "Category", "Prompt", "Field",
            "Expected", "Actual", "Deduction", "Details"]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, 1, ci, h, font_name)
    ws.row_dimensions[1].height = 18

    ri = 2
    for s in scored:
        # ── Unmatched prompt ──────────────────────────────────────────────────
        if not s["matched"]:
            vals = [s["id"], s["category"], s["prompt"],
                    "—", "—", "—", "—", "No match in standard"]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font      = Font(name=font_name, size=9, color=_UNMATCHED_FG)
                c.fill      = _hex_fill(_UNMATCHED_BG)
                c.border    = _border()
                c.alignment = Alignment(wrap_text=(ci in (3, 8)), vertical="top")
            ri += 1
            continue

        # ── Prompt text differs (informational, no deduction) ─────────────────
        if s["prompt_differs"]:
            detail = (f"Standard:  {s['expected_prompt']}\n"
                      f"Actual:    {s['prompt']}")
            vals = [s["id"], s["category"], s["prompt"],
                    "Prompt", s["expected_prompt"], s["prompt"],
                    "0 (info)", detail]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font      = Font(name=font_name, size=9, color=_INFO_FG)
                c.fill      = _hex_fill(_INFO_BG)
                c.border    = _border()
                c.alignment = Alignment(wrap_text=(ci in (3, 5, 6, 8)),
                                        vertical="top")
            ws.row_dimensions[ri].height = _row_ht(detail)
            ri += 1

        # ── Deduction rows ────────────────────────────────────────────────────
        for d in s["deductions"]:
            detail = _fmt_deduction_detail(d)

            vals = [
                s["id"], s["category"], s["prompt"],
                d["field"], d["expected"], d["actual"],
                f"-{d['deduction']}", detail or "—",
            ]
            for ci, val in enumerate(vals, 1):
                if ci == 7:   # Deduction column
                    _dat(ws, ri, ci, val, font_name,
                         bg="FFF9C4", fg="B71C1C", bold=True, align="center")
                else:
                    _dat(ws, ri, ci, val, font_name,
                         bg=_alt(ri), wrap=(ci in (3, 5, 6, 8)))
            ws.row_dimensions[ri].height = _row_ht(d["expected"], d["actual"], detail)
            ri += 1

    for ci, w in enumerate([6, 18, 42, 18, 42, 42, 12, 46], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


# ── Per-field sheets ──────────────────────────────────────────────────────────

def _write_field_sheets(wb, scored, font_name, primary, std_map, result_map):
    """
    One sheet per field in _FIELD_ORDER.
    Every matched prompt where the standard has a value for that field is listed:
      • Mismatch row  — deduction styling; Final Deduction = positive numeric value
      • Match row     — green; Final Deduction = 0
      • N/A row       — grey; field not in standard; Final Deduction = 0
    The "Final Deduction" column mirrors Deduction and is override-friendly;
    Score Card delta cells and Field Score are driven by VLOOKUP into this column.
    Fields with an Optional column (Attributes Used, Metrics Used, Other Used)
    have 8 columns; others have 7.
    """
    # Pre-index deductions by (prompt_id, field) for O(1) lookup
    ded_index: dict[tuple, dict] = {}
    for s in scored:
        for d in s["deductions"]:
            ded_index[(s["id"], d["field"])] = d

    for field in _FIELD_ORDER:
        std_key, fmt_std, result_key, fmt_actual = _FIELD_DISPLAY[field]
        opt_key     = _OPTIONAL_STD_KEY.get(field)
        has_opt_col = opt_key is not None

        rows = []
        any_expected = False
        for s in scored:
            if not s["matched"]:
                continue
            std_rec      = std_map.get(s["prompt"].strip().lower())
            has_expected = std_rec is not None and std_rec.get(std_key) is not None
            ded          = ded_index.get((s["id"], field)) if has_expected else None
            rows.append((s, std_rec, ded, has_expected))
            if has_expected:
                any_expected = True

        if not any_expected:
            continue

        wf = wb.create_sheet(_FIELD_SHEET_NAME[field])

        # Sheet title — 8 cols for Optional fields, 7 for the rest
        last_col = "H" if has_opt_col else "G"
        wf.merge_cells(f"A1:{last_col}1")
        wf["A1"] = f"Field Detail: {field}"
        wf["A1"].font      = Font(name=font_name, bold=True, size=12, color="FFFFFF")
        wf["A1"].fill      = _hex_fill(primary)
        wf["A1"].alignment = Alignment(horizontal="center")
        wf.row_dimensions[1].height = 24

        # Column headers
        if has_opt_col:
            hdrs = ["#", "Prompt", "Expected", "Optional", "Actual",
                    "Deduction", "Final Deduction", "Details"]
        else:
            hdrs = ["#", "Prompt", "Expected", "Actual",
                    "Deduction", "Final Deduction", "Details"]
        for ci, h in enumerate(hdrs, 1):
            _hdr(wf, 2, ci, h, font_name, bg="555555")

        # Column indices
        ci_actual       = 5 if has_opt_col else 4
        ci_deduct       = 6 if has_opt_col else 5
        ci_final_deduct = 7 if has_opt_col else 6   # VLOOKUP target from Score Card
        ci_details      = 8 if has_opt_col else 7

        def _opt_str(std_rec):
            if not has_opt_col or std_rec is None:
                return "—"
            items = std_rec.get(opt_key) or []
            return ", ".join(sorted(items, key=str.lower)) if items else "—"

        for ri, (s, std_rec, ded, has_expected) in enumerate(rows, 3):
            res     = result_map.get(s["id"], {})
            opt_val = _opt_str(std_rec)

            if not has_expected:
                # ── N/A row ────────────────────────────────────────────────────
                actual_str = fmt_actual(res.get(result_key))
                if has_opt_col:
                    vals = [s["id"], s["prompt"], "—", opt_val, actual_str,
                            "", 0, "N/A — not in standard"]
                else:
                    vals = [s["id"], s["prompt"], "—", actual_str,
                            "", 0, "N/A — not in standard"]
                for ci, val in enumerate(vals, 1):
                    _dat(wf, ri, ci, val, font_name,
                         bg=_NA_BG, fg=_NA_FG,
                         wrap=(ci in (2, ci_actual)),
                         align="center" if ci == ci_final_deduct else "left")
                wf.row_dimensions[ri].height = _row_ht(actual_str)

            elif ded is not None:
                # ── Mismatch row ───────────────────────────────────────────────
                detail = _fmt_deduction_detail(ded)

                if has_opt_col:
                    vals = [s["id"], s["prompt"], ded["expected"], opt_val,
                            ded["actual"], f"-{ded['deduction']}",
                            ded["deduction"], detail or "—"]
                else:
                    vals = [s["id"], s["prompt"], ded["expected"],
                            ded["actual"], f"-{ded['deduction']}",
                            ded["deduction"], detail or "—"]
                for ci, val in enumerate(vals, 1):
                    if ci in (ci_deduct, ci_final_deduct):
                        _dat(wf, ri, ci, val, font_name,
                             bg="FFF9C4", fg="B71C1C", bold=True, align="center")
                    else:
                        _dat(wf, ri, ci, val, font_name,
                             bg=_alt(ri),
                             wrap=(ci in (2, 3, ci_actual, ci_details)))
                wf.row_dimensions[ri].height = _row_ht(ded["expected"], ded["actual"], detail)

            else:
                # ── Match row ──────────────────────────────────────────────────
                expected_str = fmt_std(std_rec[std_key])
                actual_str   = fmt_actual(res.get(result_key))

                if has_opt_col:
                    vals = [s["id"], s["prompt"], expected_str, opt_val,
                            actual_str, "", 0, "Match"]
                else:
                    vals = [s["id"], s["prompt"], expected_str,
                            actual_str, "", 0, "Match"]
                for ci, val in enumerate(vals, 1):
                    _dat(wf, ri, ci, val, font_name,
                         bg=_MATCH_BG, fg=_MATCH_FG,
                         bold=(ci == ci_details),
                         wrap=(ci in (2, 3, ci_actual)),
                         align="center" if ci == ci_final_deduct else "left")
                wf.row_dimensions[ri].height = _row_ht(expected_str, actual_str)

        if has_opt_col:
            col_widths = [6, 46, 46, 30, 46, 12, 12, 46]
        else:
            col_widths = [6, 46, 46, 46, 12, 12, 46]
        for ci, w in enumerate(col_widths, 1):
            wf.column_dimensions[get_column_letter(ci)].width = w
        wf.freeze_panes = "A3"


# ── SQL issue-count formatter ─────────────────────────────────────────────────

_COUNT_LABELS: list[tuple[str, str]] = [
    ("missing_attrs",   "Missing Attrs Δ"),
    ("extra_attrs",     "Extra Attrs Δ"),
    ("missing_metrics", "Missing Metrics"),
    ("major_filters",   "Major Filter Δ"),
    ("other_major",     "Other Major Δ"),
    ("other",           "Other Δ"),
    ("minor_filters",   "Minor Filter Δ"),
    ("added_metrics",   "Added Metrics"),
]


def _fmt_counts(counts: dict | None,
                penalties: dict | None = None,
                details: dict | None = None) -> str:
    """Format issue counts as a multi-line string for the Issues cell.
    Each non-zero count shows its applied penalty, e.g. 'Major Filter Δ: 1 (-1.00)'.
    When penalty is 0 but count > 0 (waived by rule), shows '(waived)'.
    Detail lines for other_major, other, and minor_filters are appended when present.
    Returns '—' when all counts are zero; '' when counts is None."""
    if counts is None:
        return ""
    lines = []
    for k, lbl in _COUNT_LABELS:
        count = counts.get(k, 0)
        if count == 0:
            continue
        if penalties is not None:
            p = penalties.get(k, 0.0)
            suffix = f" (waived)" if p == 0.0 else f" (-{p:.2f})"
        else:
            suffix = ""
        line = f"{lbl}: {count}{suffix}"
        if details:
            detail = details.get(k, "")
            if detail:
                line += f"\n  {detail}"
        lines.append(line)
    return "\n".join(lines) if lines else "—"


# ── SQL Comparison sheet ───────────────────────────────────────────────────────

def _write_sql_sheet(wb, scored, font_name, primary, std_map, result_map):
    """
    Sheet: SQL Comparison
    Lists every matched prompt that has a sql_judgment (i.e. SQL enrichment was run).
    Colour-coded by SQL score band. Only created when at least one record has sql_judgment.
    Columns: #, Prompt, SQL Score, Issues, Gold SQL, Actual SQL, Explanation (7 cols).
    """
    rows = [(s, result_map.get(s["id"], {}))
            for s in scored
            if s.get("matched") and s.get("sql_judgment")]
    if not rows:
        return   # SQL enrichment was not run — omit sheet entirely

    ws = wb.create_sheet("SQL Comparison")

    # Sheet title
    ws.merge_cells("A1:H1")
    ws["A1"] = "SQL Comparison"
    ws["A1"].font      = Font(name=font_name, bold=True, size=12, color="FFFFFF")
    ws["A1"].fill      = _hex_fill(primary)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    # Column headers — col D is "Final SQL Score" (editable; Score Card reads from here)
    hdrs = ["#", "Prompt", "SQL Score", "Final SQL Score",
            "Issues", "Gold SQL", "Actual SQL", "Explanation"]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws, 2, ci, h, font_name, bg="555555")

    for ri, (s, res) in enumerate(rows, 3):
        j = s["sql_judgment"]

        std_rec    = std_map.get(s["prompt"].strip().lower()) or {}
        gold_sql   = format_sql(std_rec.get("sql") or "")
        actual_sql = format_sql((res.get("sqlQueries") or [""])[0])

        sql_score  = s.get("sql_score")   # float 0–5 or None (error)
        score_val  = sql_score if sql_score is not None else "err"
        issues_str = _fmt_counts(
            j.get("counts"),
            penalties=j.get("penalties"),
            details=j.get("details"),
        )

        # Final SQL Score (col D) is linked to SQL Score (col C) — user may override
        final_formula = f"=C{ri}"

        vals = [
            s["id"],
            s["prompt"],
            score_val,          # col C — SQL Score (computed)
            final_formula,      # col D — Final SQL Score (linked; override-friendly)
            issues_str,         # col E
            gold_sql,           # col F
            actual_sql,         # col G
            j.get("explanation", ""),  # col H
        ]
        for ci, val in enumerate(vals, 1):
            if ci in (3, 4):   # SQL Score / Final SQL Score — score-band colour
                score_bg, score_fg = (_score_colors(sql_score)
                                      if sql_score is not None
                                      else (_UNMATCHED_BG, _UNMATCHED_FG))
                _dat(ws, ri, ci, val, font_name,
                     bg=score_bg, fg=score_fg,
                     bold=True, align="center", valign="center")
            else:
                _dat(ws, ri, ci, val, font_name,
                     bg=_alt(ri),
                     wrap=(ci in (2, 5, 6, 7, 8)),
                     align="left", valign="top" if ci in (2, 5, 6, 7, 8) else "center")
        ws.row_dimensions[ri].height = _row_ht(
            issues_str, gold_sql, actual_sql, j.get("explanation", ""))

    for ci, w in enumerate([6, 38, 12, 12, 26, 52, 52, 46], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A3"

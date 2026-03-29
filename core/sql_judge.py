"""
core/sql_judge.py — LLM-based SQL comparison

Compares an agent's generated SQL against a gold-standard SQL using Claude.
Used as an optional enrichment step after expected-values scoring.
The sql_score (0–5) is independent of the field score and does not modify it.

─── SQL Score Categories & Deduction Rates ───────────────────────────────────
  Score starts at 5.0; each category deducts independently; floored at 0.0.

  Category          Rate    Cap   Max    Definition
  ─────────────────────────────────────────────────────────────────────────────
  missing_attrs     -1.50    2   -3.00  Gold attributes absent from actual SQL
                                        (less granular than gold); includes
                                        removed GROUP BY columns
  extra_attrs       -1.00    2   -2.00  Actual attributes not in gold SQL
                                        (more granular than gold); includes
                                        added GROUP BY columns
  missing_metrics   -1.00    3   -3.00  Gold metrics absent from actual SQL
  major_filters     -1.50    2   -3.00  Different filter column; wrong value on same
                                        column; wrong operator (= vs !=); filter
                                        present in gold but missing from actual;
                                        extra filter in actual not in gold
  other_major        -0.50    2   -1.00  HAVING differences (always); ORDER BY /
                                        LIMIT only when business question clearly
                                        requires specific sort or row limit
  added_metrics     -0.25    2   -0.50  Extra metrics in actual not in gold
  minor_filters     -0.25    2   -0.50  Cosmetic-only filter differences;
                                        waived when major_filters > 0

─── Python Pre-normalisation (applied to both SQLs before Claude sees them) ──
  1.  Strip /* block */ and -- line comments
  2.  Normalise double-quoted identifiers: strip parenthetical content,
      take display name before '__', lowercase + remove noise chars
  3.  LOWER('literal') → 'literal'
  4.  LOWER("col") / LOWER(col) → "col" / col  (column wrapper stripped)
  5.  Single-quoted string literals → lowercased  ('SuperMart Inc.' → 'supermart inc.')
  6.  COUNT(1) → COUNT(*)
  7.  <> → !=
  8.  FETCH FIRST N ROWS ONLY → LIMIT N
  9.  SELECT TOP N → SELECT … LIMIT N
  10. IN ('x') / IN (n) single-value → = 'x' / = n
  11. ISNULL / NVL / IFNULL → COALESCE
  12. Strip quoted SELECT aliases  AS "alias"
  13. If gold has no ORDER BY, strip ORDER BY from actual
  14. Collapse whitespace; strip trailing semicolon

─── Optional lists (from gold standard Excel) ────────────────────────────────
  Optional Attributes — exempt from EXTRA_ATTRS and filter penalties anywhere
  Optional Metrics    — exempt from ADDED_METRICS; Python expands with qualifier
                        prefixes (distinct / total / average / avg) before sending

─── Internal verdict states (not from Claude) ────────────────────────────────
  "ok"     — Claude responded and counts were parsed successfully
  "no_sql" — actual SQL was blank/None  (sql_score forced to 0.0)
  "error"  — API call or parse failure  (sql_score = None)
"""

import re
import datetime
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

_OUTPUT_DIR = Path(__file__).parent.parent / "output"

_API_URL = "https://api.anthropic.com/v1/messages"
_MODEL   = "claude-sonnet-4-5"


# Count-based deduction rates: (deduction_per_issue, max_issues_charged)
# Score starts at 5.0; each issue category deducts independently, floored at 0.
# Values loaded from config/scoring.yaml — edit rates there, not here.
from core import scoring_config as _sc
_SQL_DEDUCT_RATES: dict[str, tuple[float, int]] = {
    k: tuple(v)
    for k, v in _sc.load()["sql_scoring"]["deduct_rates"].items()
}


def _compute_sql_penalties(counts: dict) -> dict[str, float]:
    """
    Compute the applied penalty per category.
    Rule: minor_filters penalty is waived (set to 0) when major_filters > 0,
    since the major deduction already covers the filter difference.
    """
    skip_minor = counts.get("major_filters", 0) > 0
    penalties: dict[str, float] = {}
    for key, (rate, cap) in _SQL_DEDUCT_RATES.items():
        if key == "minor_filters" and skip_minor:
            penalties[key] = 0.0
        else:
            penalties[key] = min(counts.get(key, 0), cap) * rate
    return penalties


def _compute_sql_score(counts: dict) -> float:
    """
    Compute a 0–5 SQL score from issue counts using per-category deduction rates.
    minor_filters is waived when major_filters > 0; total is floored at 0.0.
    """
    return round(max(0.0, 5.0 - sum(_compute_sql_penalties(counts).values())), 2)


# ── SQL pre-normalisation ──────────────────────────────────────────────────────
# Applied to both gold and actual SQL before sending to Claude so that syntactic
# noise is resolved in Python rather than relying on prompt instructions.
# Add new transformations here as additional normalisation rules are identified.

_SQL_QUOTED_ID_RE  = re.compile(r'"([^"]*)"')                          # double-quoted identifier
_SQL_PAREN_CONT_RE = re.compile(r'\s*\([^)]*\)')                       # "(content)" + leading space
_SQL_NOISE_RE      = re.compile(r'[\s_()\-\[\]]')                      # noise chars to strip
_SQL_LOWER_LIT_RE  = re.compile(r"LOWER\(('[^']*')\)", re.IGNORECASE)  # LOWER('literal') → 'literal'
_SQL_LOWER_COL_RE  = re.compile(r'LOWER\s*\(\s*("(?:[^"]*)"|\w+)\s*\)', re.IGNORECASE)  # LOWER("col")/LOWER(col) → col
_SQL_STR_LIT_RE    = re.compile(r"'([^']*)'")                          # single-quoted string literal
_SQL_WHITESPACE_RE = re.compile(r'\s+')
# Additional normalisation patterns
_SQL_BLOCK_CMT_RE  = re.compile(r'/\*.*?\*/', re.DOTALL)               # /* block comment */
_SQL_LINE_CMT_RE   = re.compile(r'--[^\n]*')                           # -- line comment
_SQL_COUNT1_RE     = re.compile(r'\bCOUNT\s*\(\s*1\s*\)', re.IGNORECASE)          # COUNT(1) → COUNT(*)
_SQL_NEQ_RE        = re.compile(r'<>')                                             # <> → !=
_SQL_FETCH_RE      = re.compile(r'\bFETCH\s+FIRST\s+(\d+)\s+ROWS?\s+ONLY\b',     # FETCH FIRST N ROWS ONLY → LIMIT N
                                re.IGNORECASE)
_SQL_TOP_RE        = re.compile(r'\bSELECT\s+TOP\s+(\d+)\b', re.IGNORECASE)      # SELECT TOP N → SELECT … LIMIT N
_SQL_IN1_STR_RE    = re.compile(r'\bIN\s*\(\s*(\'[^\']*\')\s*\)', re.IGNORECASE)  # IN ('x') → = 'x'
_SQL_IN1_NUM_RE    = re.compile(r'\bIN\s*\(\s*(\d+(?:\.\d+)?)\s*\)', re.IGNORECASE) # IN (n) → = n
_SQL_ISNULL_RE     = re.compile(r'\bISNULL\s*\(', re.IGNORECASE)                  # ISNULL( → COALESCE(
_SQL_NVL_RE        = re.compile(r'\bNVL\s*\(', re.IGNORECASE)                     # NVL( → COALESCE(
_SQL_IFNULL_RE     = re.compile(r'\bIFNULL\s*\(', re.IGNORECASE)                  # IFNULL( → COALESCE(
# AS alias stripping: matches  expr AS "alias"  at a column boundary.
# Only quoted aliases are stripped (CAST uses unquoted type names).
# Aliases that appear more than once in the SQL are preserved — they are
# referenced downstream (ORDER BY, HAVING, CTE body) and stripping them
# would corrupt the query structure Claude sees.
_SQL_ALIAS_RE      = re.compile(
    r'\s+AS\s+"([^"]*)"'
    r'(?=\s*(?:,|\)|\bFROM\b|\bWHERE\b|\bGROUP\b|\bORDER\b|\bHAVING\b|\bLIMIT\b|\bUNION\b|$))',
    re.IGNORECASE,
)


def _strip_aliases_safe(sql: str) -> str:
    """Strip AS "alias" only when the alias is not referenced elsewhere in the SQL.
    Aliases appearing more than once are kept — they are referenced downstream
    (e.g. in ORDER BY, HAVING, or a CTE body) and must not be removed."""
    def _replacer(m: re.Match) -> str:
        alias = m.group(1)
        # Count all occurrences of "alias" as a quoted identifier in the full SQL
        occurrences = len(re.findall(
            r'"' + re.escape(alias) + r'"', sql, re.IGNORECASE
        ))
        return '' if occurrences <= 1 else m.group(0)
    return _SQL_ALIAS_RE.sub(_replacer, sql)


def _norm_identifier(name: str) -> str:
    """Normalise a SQL identifier (column or table name).
      1. Strip parenthetical content  — e.g. '(customer level 6 planning account name)'
      2. Keep display name before first '__'  — MicroStrategy DisplayName__InternalName pattern
      3. Lowercase + remove spaces, underscores, hyphens, brackets
    """
    name = _SQL_PAREN_CONT_RE.sub('', name)
    if '__' in name:
        name = name.split('__', 1)[0]
    return _SQL_NOISE_RE.sub('', name.lower())


def _normalise_sql(sql: str) -> str:
    """Pre-normalise a SQL string before sending to Claude for comparison.
    Both gold and actual SQL are normalised so Claude receives comparable text.
    Transformations applied in order:
      1.  Strip block comments        /* ... */
      2.  Strip line comments         -- ...
      3.  Double-quoted identifiers   → _norm_identifier() (case, noise, __ , parens)
      4.  LOWER('literal')            → 'literal'
      5.  LOWER("col") / LOWER(col)  → "col" / col  (column wrapper only; literals already gone)
      6.  Single-quoted string literals → lowercased  ('SuperMart Inc.' → 'supermart inc.')
      7.  COUNT(1)                    → COUNT(*)
      8.  <> (not-equal)             → !=
      9.  FETCH FIRST N ROWS ONLY    → LIMIT N
      10. SELECT TOP N               → SELECT … LIMIT N  (N appended at end)
      11. IN ('x') / IN (n)          → = 'x' / = n  (single-value IN only)
      12. ISNULL / NVL / IFNULL      → COALESCE
      13. Strip quoted SELECT aliases  AS "alias" (CAST uses unquoted type names)
      14. Collapse whitespace + strip trailing semicolon
    """
    # 1–2: comments
    sql = _SQL_BLOCK_CMT_RE.sub(' ', sql)
    sql = _SQL_LINE_CMT_RE.sub(' ', sql)
    # 3: identifier normalisation
    sql = _SQL_QUOTED_ID_RE.sub(lambda m: f'"{_norm_identifier(m.group(1))}"', sql)
    # 4: LOWER('literal') → 'literal'
    sql = _SQL_LOWER_LIT_RE.sub(r'\1', sql)
    # 5: LOWER("col") / LOWER(col) → "col" / col  (literals already stripped in step 4)
    sql = _SQL_LOWER_COL_RE.sub(r'\1', sql)
    # 6: lowercase all single-quoted string literals for case-insensitive value comparison
    sql = _SQL_STR_LIT_RE.sub(lambda m: f"'{m.group(1).lower()}'", sql)
    # 7: COUNT(1) → COUNT(*)
    sql = _SQL_COUNT1_RE.sub('COUNT(*)', sql)
    # 8: <> → !=
    sql = _SQL_NEQ_RE.sub('!=', sql)
    # 9: FETCH FIRST N ROWS ONLY → LIMIT N
    sql = _SQL_FETCH_RE.sub(lambda m: f'LIMIT {m.group(1)}', sql)
    # 10: SELECT TOP N → SELECT (append LIMIT N at end)
    top_m = _SQL_TOP_RE.search(sql)
    if top_m:
        n = top_m.group(1)
        sql = _SQL_TOP_RE.sub('SELECT', sql)
        sql = sql.rstrip().rstrip(';').strip() + f' LIMIT {n}'
    # 11: single-value IN → equality
    sql = _SQL_IN1_STR_RE.sub(r"= \1", sql)
    sql = _SQL_IN1_NUM_RE.sub(r"= \1", sql)
    # 12: null-coalescing dialect variants → COALESCE
    sql = _SQL_ISNULL_RE.sub('COALESCE(', sql)
    sql = _SQL_NVL_RE.sub('COALESCE(', sql)
    sql = _SQL_IFNULL_RE.sub('COALESCE(', sql)
    # 13: strip quoted SELECT aliases only when not referenced downstream
    sql = _strip_aliases_safe(sql)
    # 14: collapse whitespace + strip trailing semicolon
    sql = _SQL_WHITESPACE_RE.sub(' ', sql).strip().rstrip(';').strip()
    return sql


def _build_identifier_map(raw_sql: str) -> dict[str, str]:
    """Build a normalised-key → human-readable display name mapping from raw SQL.

    For each double-quoted identifier found:
      1. Strip parenthetical content  (same as _norm_identifier step 1)
      2. Take the part before the first '__'  (same as step 2)
      3. The result is the display name — stored as-is (preserving original casing/spacing)
      4. The normalised key is derived by applying step 3 of _norm_identifier to the display name

    This lets judge_sql() tell Claude which original names correspond to each
    normalised identifier so Claude uses them in EXPLANATION and detail fields.
    """
    mapping: dict[str, str] = {}
    for m in _SQL_QUOTED_ID_RE.finditer(raw_sql):
        raw      = m.group(1)
        display  = _SQL_PAREN_CONT_RE.sub('', raw).strip()   # step 1: strip parens
        if '__' in display:
            display = display.split('__', 1)[0].strip()      # step 2: before first __
        norm_key = _SQL_NOISE_RE.sub('', display.lower())    # step 3: derive key
        if norm_key and norm_key not in mapping:
            mapping[norm_key] = display
    return mapping


# System prompt loaded from config/scoring.yaml — edit it there, not here.
_SYSTEM: str = _sc.load()["sql_scoring"]["system_prompt"].rstrip()

# Leading qualifier words that may prefix a metric alias without changing its identity.
# Used in the optional metrics matching instruction sent to Claude (SQL comparison path only).
# After normalisation these are lowercase with no spaces: "distinct", "total", "average", "avg".
_METRIC_QUALIFIERS: frozenset[str] = frozenset({"distinct", "total", "average", "avg"})

_EMPTY_COUNTS: dict[str, int] = {
    "missing_attrs":   0,
    "extra_attrs":     0,
    "missing_metrics": 0,
    "major_filters":   0,
    "other_major":     0,
    "other":           0,
    "minor_filters":   0,
    "added_metrics":   0,
}

_COUNT_PATTERNS: list[tuple[str, str]] = [
    ("missing_attrs",   r"MISSING_ATTRS\s*:\s*(\d+)"),
    ("extra_attrs",     r"EXTRA_ATTRS\s*:\s*(\d+)"),
    ("missing_metrics", r"MISSING_METRICS\s*:\s*(\d+)"),
    ("major_filters",   r"MAJOR_FILTERS\s*:\s*(\d+)"),
    ("other_major",     r"OTHER_MAJOR\s*:\s*(\d+)"),
    ("other",           r"\bOTHER\b\s*:\s*(\d+)"),
    ("minor_filters",   r"MINOR_FILTERS\s*:\s*(\d+)"),
    ("added_metrics",   r"ADDED_METRICS\s*:\s*(\d+)"),
]

_DETAIL_PATTERNS: list[tuple[str, str]] = [
    ("missing_attrs",   r"MISSING_ATTRS_DETAIL\s*:\s*(.+)"),
    ("extra_attrs",     r"EXTRA_ATTRS_DETAIL\s*:\s*(.+)"),
    ("missing_metrics", r"MISSING_METRICS_DETAIL\s*:\s*(.+)"),
    ("major_filters",   r"MAJOR_FILTERS_DETAIL\s*:\s*(.+)"),
    ("other_major",     r"OTHER_MAJOR_DETAIL\s*:\s*(.+)"),
    ("other",           r"OTHER_DETAIL\s*:\s*(.+)"),
    ("minor_filters",   r"MINOR_FILTERS_DETAIL\s*:\s*(.+)"),
    ("added_metrics",   r"ADDED_METRICS_DETAIL\s*:\s*(.+)"),
]


# ── Internal helpers ──────────────────────────────────────────────────────────

def _call_claude(user: str, api_key: str, system: str) -> str:
    resp = requests.post(
        _API_URL,
        headers={
            "x-api-key":         api_key,
            "anthropic-version": "2023-06-01",
            "content-type":      "application/json",
        },
        json={
            "model":      _MODEL,
            "max_tokens": 1024,
            "system":     system,
            "messages":   [{"role": "user", "content": user}],
        },
        timeout=60,
        verify=False,
    )
    resp.raise_for_status()
    return resp.json()["content"][0]["text"].strip()


def _parse_response(raw: str) -> tuple[str, dict | None, dict]:
    """
    Extract explanation, issue counts, and category details from Claude's response.
    Returns (explanation, counts, details).
    counts is None when the response cannot be parsed.
    details keys: all 8 categories — empty string when absent or Claude wrote 'none'.
    """
    m_e = re.search(r"EXPLANATION\s*:\s*(.+)", raw, re.IGNORECASE | re.DOTALL)
    explanation = m_e.group(1).strip().split("\n")[0] if m_e else ""
    counts = {
        key: int(m.group(1)) if (m := re.search(pat, raw, re.IGNORECASE)) else 0
        for key, pat in _COUNT_PATTERNS
    }
    # Require at least one count pattern to match; otherwise treat as parse failure
    if not any(re.search(pat, raw, re.IGNORECASE) for _, pat in _COUNT_PATTERNS):
        return explanation, None, {}
    details: dict[str, str] = {}
    for key, pat in _DETAIL_PATTERNS:
        m = re.search(pat, raw, re.IGNORECASE)
        if m:
            val = m.group(1).strip().split("\n")[0].strip()
            details[key] = "" if val.lower() == "none" else val
        else:
            details[key] = ""
    return explanation, counts, details


# ── Public API ────────────────────────────────────────────────────────────────

def judge_sql(gold_sql: str, actual_sql: str | None,
              prompt_text: str, api_key: str,
              agent_instructions: str = "",
              optional_attrs: list[str] | None = None,
              optional_metrics: list[str] | None = None) -> dict:
    """
    Compare actual_sql against gold_sql for the given business question.

    agent_instructions: optional text from the agent's CI file; injected into
    the system prompt so Claude can distinguish expected agent behavior (e.g.
    metrics the agent always includes by convention) from genuine errors.

    optional_attrs: attributes that are permitted extras — if present in the
    actual SQL but absent from the gold SQL, they must NOT be counted as
    EXTRA_ATTRS (penalty = 0).

    optional_metrics: metrics that are permitted extras — if present in the
    actual SQL but absent from the gold SQL, they must NOT be counted as
    ADDED_METRICS (penalty = 0).

    Returns a dict:
      { verdict, explanation, sql_score, counts, penalties, details }

    Internal verdict states (not from Claude):
      "ok"     — Claude responded and counts were parsed successfully
      "no_sql" — actual_sql was blank/None (sql_score forced to 0.0)
      "error"  — API call failed or response could not be parsed (sql_score = None)
    """
    if not actual_sql or not actual_sql.strip():
        return {
            "verdict":     "no_sql",   # internal flag — not from Claude
            "explanation": "Agent produced no SQL.",
            "sql_score":   0.0,
            "counts":      dict(_EMPTY_COUNTS),
            "penalties":   {k: 0.0 for k in _SQL_DEDUCT_RATES},
            "details":     {},
        }

    # Build system prompt — append agent instructions as behavioral context if provided
    system = _SYSTEM
    if agent_instructions.strip():
        system += (
            "\n\nAgent behavior context — the following are the agent's own instructions "
            "that govern how it generates SQL. Use this to understand whether a difference "
            "is a real error or expected agent behavior. Pay particular attention to:\n"
            "  • Metrics the agent always includes by convention (do not penalise as ADDED_METRICS)\n"
            "  • Filter column hierarchies defined — if the actual SQL uses a different column\n"
            "    from the same hierarchy as the gold SQL, do not penalise as MAJOR_FILTERS\n"
            "    when the agent instructions indicate this substitution is acceptable or expected\n"
            "  • Default output format and sort order preferences\n"
            "Do not penalise differences that are consistent with these instructions:\n\n"
            + agent_instructions.strip()
        )

    # Pre-normalise both SQLs — resolves identifier and LOWER() differences in Python
    gold_sql_n   = _normalise_sql(gold_sql)
    actual_sql_n = _normalise_sql(actual_sql)

    # If the gold SQL has no ORDER BY, strip any ORDER BY from the actual SQL before
    # sending to Claude so the difference is invisible and cannot be penalised.
    _ORDER_BY_RE = re.compile(r'\bORDER\s+BY\b.*', re.IGNORECASE | re.DOTALL)
    if not _ORDER_BY_RE.search(gold_sql_n):
        actual_sql_n = _ORDER_BY_RE.sub('', actual_sql_n).strip()

    # Build normalised→display map so Claude can use original names in free-text fields.
    # Process actual first, then gold — gold overrides to get its (usually better) casing.
    id_map = _build_identifier_map(actual_sql or "")
    id_map.update(_build_identifier_map(gold_sql))

    # Build user message — append optional attrs/metrics notes when provided
    # Normalise names to match the already-normalised SQL identifiers
    optional_note = ""
    norm_opt_attrs = [_norm_identifier(a) for a in (optional_attrs or [])]
    if norm_opt_attrs:
        attrs_csv = ", ".join(norm_opt_attrs)
        optional_note += (
            f"\n\nOptional attributes for this question (pre-normalised): {attrs_csv}\n"
            "These are permitted extras in SELECT, GROUP BY, and ORDER BY — if they appear\n"
            "in those clauses in the actual SQL but not in the gold SQL, do NOT count them\n"
            "as EXTRA_ATTRS or raise them as grouping/ordering issues (no penalty there).\n"
            "IMPORTANT: this exemption does NOT apply to WHERE or HAVING. If an optional\n"
            "attribute appears as a filter condition in WHERE or HAVING when the gold SQL\n"
            "does not filter on it, count it as MAJOR_FILTERS — it changes which rows are returned.\n"
            "Match by exact normalised name only."
        )

    norm_opt_metrics = [_norm_identifier(m) for m in (optional_metrics or [])]
    if norm_opt_metrics:
        # Expand with qualifier-prefixed variants so Claude only needs exact matching.
        # e.g. "promocount" → also "distinctpromocount", "totalpromocount", etc.
        expanded_opt_metrics: list[str] = []
        for base in norm_opt_metrics:
            expanded_opt_metrics.append(base)
            for q in _METRIC_QUALIFIERS:
                expanded_opt_metrics.append(q + base)
        metrics_csv = ", ".join(expanded_opt_metrics)
        optional_note += (
            f"\n\nOptional metrics for this question (pre-normalised, including qualifier variants): {metrics_csv}\n"
            "These are fully permitted extras — if they appear anywhere in the actual SQL\n"
            "(SELECT, aggregations, ORDER BY, or any other clause) but not in the gold SQL,\n"
            "do NOT count them as ADDED_METRICS or raise them as issues anywhere else\n"
            "(they carry no penalty in any category). Match by exact normalised name only."
        )

    # Identifier reference: only include entries where the display name differs from the key
    id_ref_lines = [
        f"  {key} → \"{disp}\""
        for key, disp in sorted(id_map.items())
        if key != disp.lower().replace(' ', '').replace('_', '')
           .replace('-', '').replace('(', '').replace(')', '').replace('[', '').replace(']', '')
    ]
    id_ref_note = ""
    if id_ref_lines:
        id_ref_note = (
            "\n\nIdentifier reference (normalised key → original display name):\n"
            + "\n".join(id_ref_lines)
            + "\nIn EXPLANATION, OTHER_MAJOR_DETAIL, and MINOR_FILTERS_DETAIL, always refer to "
            "columns, tables, and metrics using the original display names from this reference, "
            "not the normalised keys."
        )

    user = (
        f'Business question: "{prompt_text}"\n\n'
        f"Gold-standard SQL:\n```sql\n{gold_sql_n}\n```\n\n"
        f"Actual SQL:\n```sql\n{actual_sql_n}\n```\n\n"
        "Compare the two queries and respond with counts, details, and EXPLANATION."
        + optional_note
        + id_ref_note
    )

    try:
        raw                          = _call_claude(user, api_key, system)
        explanation, counts, details = _parse_response(raw)
        verdict                      = "ok"
    except Exception as exc:
        explanation = str(exc)[:200]
        counts      = None
        details     = {}
        verdict     = "error"

    # Compute sql_score and per-category penalties from counts
    if counts is not None:
        penalties = _compute_sql_penalties(counts)
        sql_score = round(max(0.0, 5.0 - sum(penalties.values())), 2)
    else:
        penalties = None
        sql_score = None   # parse/API error — unscored

    return {
        "verdict":     verdict,     # "ok" | "no_sql" | "error" — internal only, not from Claude
        "explanation": explanation,
        "sql_score":   sql_score,
        "counts":      counts,
        "penalties":   penalties,
        "details":     details,
    }


def _write_sql_debug_xlsx(rows: list[dict]) -> Path:
    """Write a debug Excel file showing the normalised inputs sent to Claude per prompt.

    Columns: #, Prompt, Attributes Used, Optional Attrs (raw), Optional Attrs (normalised),
             Metrics Used, Gold SQL (normalised), Actual SQL (normalised).

    Returns the path of the written file.
    """
    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = _OUTPUT_DIR / f"sql_debug_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SQL Debug"

    headers = [
        "#",
        "Prompt",
        "Attributes Used",
        "Optional Attrs (raw)",
        "Optional Attrs (normalised)",
        "Metrics Used",
        "Optional Metrics (raw)",
        "Optional Metrics (normalised)",
        "WHERE Tokens (actual)",
        "Gold SQL (normalised)",
        "Actual SQL (normalised)",
    ]

    # Header style
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="36454F")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align

    # Data rows
    wrap = Alignment(vertical="top", wrap_text=True)
    top  = Alignment(vertical="top", wrap_text=False)
    for r_idx, row in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1,  value=row["num"]).alignment = top
        ws.cell(row=r_idx, column=2,  value=row["prompt"]).alignment = wrap
        ws.cell(row=r_idx, column=3,  value=row["attrs_used"]).alignment = wrap
        ws.cell(row=r_idx, column=4,  value=row["optional_attrs_raw"]).alignment = wrap
        ws.cell(row=r_idx, column=5,  value=row["optional_attrs_norm"]).alignment = wrap
        ws.cell(row=r_idx, column=6,  value=row["metrics_used"]).alignment = wrap
        ws.cell(row=r_idx, column=7,  value=row["optional_metrics_raw"]).alignment = wrap
        ws.cell(row=r_idx, column=8,  value=row["optional_metrics_norm"]).alignment = wrap
        ws.cell(row=r_idx, column=9,  value=row["where_tokens"]).alignment = wrap
        ws.cell(row=r_idx, column=10, value=row["gold_sql_norm"]).alignment = wrap
        ws.cell(row=r_idx, column=11, value=row["actual_sql_norm"]).alignment = wrap

    # Column widths: fixed for SQL columns, auto-capped for others
    col_widths = [4, 40, 35, 35, 35, 35, 35, 35, 40, 60, 60]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    wb.save(out_path)
    return out_path


def enrich_with_sql(scored: list[dict], envelope: dict,
                    standard: list[dict], api_key: str,
                    agent_instructions: str = "") -> int:
    """
    Enrich scored records in-place with SQL judgments.

    For each matched scored record where the standard has a gold SQL:
      • Calls judge_sql() to compare gold SQL vs the agent's actual SQL
      • Adds  rec["sql_judgment"] = { verdict, explanation, deduction, sql_score,
                                      counts, penalties, details }
      • Adds  rec["sql_score"]    = 0–5 float (None when verdict is "error")

    agent_instructions: optional text from the agent's CI file; passed through
    to every judge_sql() call so Claude can contextualise expected behavior.

    NOTE: rec["score"] (field score) is intentionally NOT modified.
    SQL quality is tracked separately on its own 0–5 scale.

    The actual SQL is read from result["sqlQueries"][0] (JSON results format).
    Returns the number of SQL comparisons attempted.
    """
    std_map    = {r["prompt"].strip().lower(): r for r in standard}
    result_map = {r.get("id"): r for r in envelope.get("results", [])}
    count      = 0
    debug_rows: list[dict] = []

    for rec in scored:
        if not rec.get("matched"):
            continue
        std = std_map.get(rec["prompt"].strip().lower())
        if std is None or not std.get("sql"):
            continue   # no gold SQL for this prompt — skip silently

        gold_sql         = std["sql"]
        result           = result_map.get(rec["id"], {})
        actual_sql       = (result.get("sqlQueries") or [""])[0]
        optional_attrs   = std.get("optional_attrs") or []
        optional_metrics = std.get("optional_metrics") or []

        # Capture normalised inputs for debug file
        gold_sql_n         = _normalise_sql(gold_sql)
        actual_sql_n       = _normalise_sql(actual_sql or "")
        norm_opt_attrs     = [_norm_identifier(a) for a in optional_attrs]
        norm_opt_metrics   = [_norm_identifier(m) for m in optional_metrics]
        where_tokens = result.get("whereClauseTokens") or []
        debug_rows.append({
            "num":                  len(debug_rows) + 1,
            "prompt":               rec["prompt"],
            "attrs_used":           ", ".join(result.get("attributesUsed") or []),
            "optional_attrs_raw":   ", ".join(optional_attrs),
            "optional_attrs_norm":  ", ".join(norm_opt_attrs),
            "metrics_used":         ", ".join(result.get("metricsUsed") or []),
            "optional_metrics_raw":  ", ".join(optional_metrics),
            "optional_metrics_norm": ", ".join(norm_opt_metrics),
            "where_tokens":         "\n".join(where_tokens) if where_tokens else "(none)",
            "gold_sql_norm":        gold_sql_n,
            "actual_sql_norm":      actual_sql_n,
        })

        judgment            = judge_sql(gold_sql, actual_sql, rec["prompt"],
                                        api_key,
                                        agent_instructions=agent_instructions,
                                        optional_attrs=optional_attrs,
                                        optional_metrics=optional_metrics)
        rec["sql_judgment"] = judgment
        rec["sql_score"]    = judgment.get("sql_score")   # None for "error" verdict
        count += 1

    # Write debug file whenever at least one comparison was run (if enabled in settings)
    if debug_rows:
        try:
            from config.settings import SQL_DEBUG as _sql_debug
        except Exception:
            _sql_debug = False
        if _sql_debug:
            try:
                debug_path = _write_sql_debug_xlsx(debug_rows)
                print(f"  [debug] SQL normalisation inputs: {debug_path}")
            except Exception as exc:
                print(f"  [debug] Could not write SQL debug file: {exc}")

    return count
